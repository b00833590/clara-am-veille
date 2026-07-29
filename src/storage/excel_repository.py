from pathlib import Path

from openpyxl import Workbook, load_workbook

from src.models import JobPosting
from src.storage.repository import JobRepository

SHEET_OFF_TOPIC = "Hors sujet"
SHEET_AM = "AM — Prioritaire"
SHEET_FINANCE = "Finance — Autres opportunités"

# Pre-dates the introduction of the N ("Hors sujet") classification category —
# workbooks created before that change hold their off-topic rows under this
# older name instead.
LEGACY_OFF_TOPIC_SHEET_NAME = "Non classifiées"

HEADERS = [
    "ID",
    "Entreprise",
    "Titre du poste",
    "Catégorie (A/B)",
    "À vérifier",
    "Raison (classification)",
    "Date de détection",
    "Date de début du stage",
    "Langue de l'annonce",
    "Lieu",
    "Priorité géo",
    "Lien vers l'offre",
    "Statut",
    "Lettre de motivation générée",
]


class ExcelJobRepository(JobRepository):
    def __init__(self, path: Path):
        self._path = Path(path)
        if not self._path.exists():
            self._create_workbook()
        else:
            self._migrate_workbook()

    def _create_workbook(self) -> None:
        workbook = Workbook()
        workbook.remove(workbook.active)
        for sheet_name in (SHEET_OFF_TOPIC, SHEET_AM, SHEET_FINANCE):
            sheet = workbook.create_sheet(sheet_name)
            sheet.append(HEADERS)
        workbook.save(self._path)

    def _migrate_workbook(self) -> None:
        """Brings a workbook created by an older version of the schema up to
        date: renames the legacy off-topic sheet if present (keeping its
        rows), then creates any of the three expected sheets still missing.
        """
        workbook = load_workbook(self._path)
        changed = False

        if SHEET_OFF_TOPIC not in workbook.sheetnames and LEGACY_OFF_TOPIC_SHEET_NAME in workbook.sheetnames:
            workbook[LEGACY_OFF_TOPIC_SHEET_NAME].title = SHEET_OFF_TOPIC
            changed = True

        for sheet_name in (SHEET_OFF_TOPIC, SHEET_AM, SHEET_FINANCE):
            if sheet_name not in workbook.sheetnames:
                sheet = workbook.create_sheet(sheet_name)
                sheet.append(HEADERS)
                changed = True
                continue

            sheet = workbook[sheet_name]
            current_header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if list(current_header_row or ()) != HEADERS:
                # A stale header row (from before a column was added/reordered)
                # doesn't just mislabel columns when zipped against newer,
                # wider data rows — duplicate padding-None header slots
                # collide as dict keys and silently drop data (confirmed:
                # URL and status were being discarded this way). Rewriting
                # the header row in place is safe: data cells are untouched.
                for col_index, header in enumerate(HEADERS, start=1):
                    sheet.cell(row=1, column=col_index, value=header)
                changed = True

        if changed:
            workbook.save(self._path)

    def _sheet_for_category(self, category: str | None) -> str:
        if category == "A":
            return SHEET_AM
        if category == "B":
            return SHEET_FINANCE
        return SHEET_OFF_TOPIC

    def exists(self, stable_id: str) -> bool:
        workbook = load_workbook(self._path)
        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if row and row[0] == stable_id:
                    return True
        return False

    def add(self, posting: JobPosting) -> None:
        stable_id = posting.stable_id()
        if self.exists(stable_id):
            return

        workbook = load_workbook(self._path)
        sheet = workbook[self._sheet_for_category(posting.category)]
        sheet.append(
            [
                stable_id,
                posting.company,
                posting.title,
                posting.category or "",
                "Oui" if posting.to_verify else "",
                posting.classification_reason,
                posting.detected_at.strftime("%Y-%m-%d %H:%M"),
                posting.start_date or "",
                posting.language or "",
                posting.location or "",
                posting.location_priority,
                posting.url or "",
                posting.status,
                posting.cover_letter_link or "",
            ]
        )
        workbook.save(self._path)

    def update_letter(self, stable_id: str, letter_text: str) -> None:
        workbook = load_workbook(self._path)
        letter_column_index = HEADERS.index("Lettre de motivation générée")

        for sheet_name in workbook.sheetnames:
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows(min_row=2):
                if row[0].value == stable_id:
                    row[letter_column_index].value = letter_text
                    workbook.save(self._path)
                    return

    def _read_sheet_rows(self, sheet_name: str) -> list[dict]:
        workbook = load_workbook(self._path)
        sheet = workbook[sheet_name]
        rows_iter = sheet.iter_rows(min_row=1, max_row=1, values_only=True)
        headers = next(rows_iter)
        rows = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            if row[0] is None:
                continue
            rows.append(dict(zip(headers, row)))
        return rows
