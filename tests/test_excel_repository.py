from openpyxl import Workbook

from src.models import JobPosting
from src.storage.excel_repository import ExcelJobRepository, HEADERS, SHEET_AM, SHEET_FINANCE, SHEET_OFF_TOPIC


def make_posting(**overrides):
    defaults = dict(
        company="Sycomore Asset Management",
        title="Stage Analyste Financier",
        url="https://careers.smartrecruiters.com/SycomoreAssetManagement/job/123",
        description="Stage au sein de l'équipe gestion...",
    )
    defaults.update(overrides)
    return JobPosting(**defaults)


def test_add_creates_workbook_file(tmp_path):
    repo = ExcelJobRepository(tmp_path / "offres.xlsx")

    repo.add(make_posting())

    assert (tmp_path / "offres.xlsx").exists()


def test_add_persists_location_and_location_priority(tmp_path):
    repo = ExcelJobRepository(tmp_path / "offres.xlsx")
    repo.add(make_posting(category="A", location="Paris, France", location_priority=1))

    rows = repo._read_sheet_rows(SHEET_AM)

    assert rows[0]["Lieu"] == "Paris, France"
    assert rows[0]["Priorité géo"] == 1


def test_add_persists_classification_reason(tmp_path):
    repo = ExcelJobRepository(tmp_path / "offres.xlsx")
    repo.add(make_posting(category="A", to_verify=True, classification_reason="Ambigu entre gestion privée et corporate banking"))

    rows = repo._read_sheet_rows(SHEET_AM)

    assert rows[0]["Raison (classification)"] == "Ambigu entre gestion privée et corporate banking"


def test_add_routes_category_a_to_am_prioritaire_sheet(tmp_path):
    repo = ExcelJobRepository(tmp_path / "offres.xlsx")
    repo.add(make_posting(category="A"))

    rows = repo._read_sheet_rows(SHEET_AM)

    assert len(rows) == 1
    assert rows[0]["Entreprise"] == "Sycomore Asset Management"


def test_add_routes_category_b_to_finance_autres_sheet(tmp_path):
    repo = ExcelJobRepository(tmp_path / "offres.xlsx")
    repo.add(make_posting(category="B"))

    rows = repo._read_sheet_rows(SHEET_FINANCE)

    assert len(rows) == 1


def test_add_routes_unclassified_posting_to_holding_sheet(tmp_path):
    repo = ExcelJobRepository(tmp_path / "offres.xlsx")
    repo.add(make_posting(category=None))

    rows = repo._read_sheet_rows(SHEET_OFF_TOPIC)

    assert len(rows) == 1


def test_exists_returns_false_for_unknown_stable_id(tmp_path):
    repo = ExcelJobRepository(tmp_path / "offres.xlsx")

    assert repo.exists("unknown-id") is False


def test_exists_returns_true_after_add_regardless_of_sheet(tmp_path):
    repo = ExcelJobRepository(tmp_path / "offres.xlsx")
    posting = make_posting(category="A")
    repo.add(posting)

    assert repo.exists(posting.stable_id()) is True


def test_exists_persists_when_reopening_repository_from_disk(tmp_path):
    path = tmp_path / "offres.xlsx"
    posting = make_posting(category="B")
    ExcelJobRepository(path).add(posting)

    reopened = ExcelJobRepository(path)

    assert reopened.exists(posting.stable_id()) is True


def test_add_is_idempotent_for_same_stable_id(tmp_path):
    repo = ExcelJobRepository(tmp_path / "offres.xlsx")
    posting = make_posting(category="A")

    repo.add(posting)
    repo.add(posting)

    rows = repo._read_sheet_rows(SHEET_AM)
    assert len(rows) == 1


def test_update_letter_sets_column_for_matching_stable_id(tmp_path):
    repo = ExcelJobRepository(tmp_path / "offres.xlsx")
    posting = make_posting(category="A")
    repo.add(posting)

    repo.update_letter(posting.stable_id(), "Madame, Monsieur, corps de la lettre...")

    rows = repo._read_sheet_rows(SHEET_AM)
    assert rows[0]["Lettre de motivation générée"] == "Madame, Monsieur, corps de la lettre..."


def test_update_letter_does_nothing_for_unknown_stable_id(tmp_path):
    repo = ExcelJobRepository(tmp_path / "offres.xlsx")
    repo.add(make_posting(category="A"))

    repo.update_letter("unknown-id", "texte")

    rows = repo._read_sheet_rows(SHEET_AM)
    assert not rows[0]["Lettre de motivation générée"]


def test_add_category_n_does_not_crash_on_workbook_missing_off_topic_sheet(tmp_path):
    path = tmp_path / "offres.xlsx"
    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet_name in (SHEET_AM, SHEET_FINANCE):
        sheet = workbook.create_sheet(sheet_name)
        sheet.append(HEADERS)
    workbook.save(path)

    repo = ExcelJobRepository(path)
    repo.add(make_posting(category="N"))

    rows = repo._read_sheet_rows(SHEET_OFF_TOPIC)
    assert len(rows) == 1


def test_reopening_workbook_updates_stale_header_row_so_new_rows_read_correctly(tmp_path):
    path = tmp_path / "offres.xlsx"
    legacy_headers = [
        "ID",
        "Entreprise",
        "Titre du poste",
        "Catégorie (A/B)",
        "À vérifier",
        "Date de détection",
        "Date de début du stage",
        "Langue de l'annonce",
        "Lien vers l'offre",
        "Statut",
        "Lettre de motivation générée",
    ]
    workbook = Workbook()
    workbook.remove(workbook.active)
    for sheet_name in (SHEET_OFF_TOPIC, SHEET_AM, SHEET_FINANCE):
        sheet = workbook.create_sheet(sheet_name)
        sheet.append(legacy_headers)
    workbook.save(path)

    repo = ExcelJobRepository(path)
    repo.add(
        make_posting(
            category="A",
            url="https://careers.example.com/job/1",
            location="Paris, France",
            location_priority=1,
            classification_reason="Analyse de portefeuille",
        )
    )

    rows = repo._read_sheet_rows(SHEET_AM)
    assert rows[0]["Raison (classification)"] == "Analyse de portefeuille"
    assert rows[0]["Lieu"] == "Paris, France"
    assert rows[0]["Priorité géo"] == 1
    assert rows[0]["Lien vers l'offre"] == "https://careers.example.com/job/1"
    assert rows[0]["Statut"] == "Nouvelle"


def test_reopening_workbook_migrates_legacy_off_topic_sheet_name_and_keeps_its_rows(tmp_path):
    path = tmp_path / "offres.xlsx"
    legacy_name = "Non classifiées"
    workbook = Workbook()
    workbook.remove(workbook.active)
    legacy_sheet = workbook.create_sheet(legacy_name)
    legacy_sheet.append(HEADERS)
    legacy_sheet.append(["legacy-id", "Ancienne Société", "Ancien Poste"])
    for sheet_name in (SHEET_AM, SHEET_FINANCE):
        sheet = workbook.create_sheet(sheet_name)
        sheet.append(HEADERS)
    workbook.save(path)

    repo = ExcelJobRepository(path)

    assert repo.exists("legacy-id") is True
    rows = repo._read_sheet_rows(SHEET_OFF_TOPIC)
    assert rows[0]["Entreprise"] == "Ancienne Société"
