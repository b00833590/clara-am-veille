from openpyxl import load_workbook

from src.models import JobPosting
from src.storage.excel_export import export_to_excel
from src.storage.sqlite_repository import SQLiteJobRepository


def make_posting(**overrides):
    defaults = dict(
        company="Sycomore Asset Management",
        title="Stage Analyste Financier",
        url="https://careers.smartrecruiters.com/SycomoreAssetManagement/job/123",
        description="Stage au sein de l'équipe gestion...",
    )
    defaults.update(overrides)
    return JobPosting(**defaults)


def test_export_creates_xlsx_with_three_sheets(tmp_path):
    repo = SQLiteJobRepository(tmp_path / "postings.db")
    xlsx_path = tmp_path / "export.xlsx"

    export_to_excel(repo, xlsx_path)

    workbook = load_workbook(xlsx_path)
    assert set(workbook.sheetnames) == {"AM — Prioritaire", "Finance — Autres opportunités", "Hors sujet"}


def test_export_routes_postings_to_the_right_sheet_by_category(tmp_path):
    repo = SQLiteJobRepository(tmp_path / "postings.db")
    repo.add(make_posting(url="https://example.com/a", category="A"))
    repo.add(make_posting(url="https://example.com/b", category="B"))
    repo.add(make_posting(url="https://example.com/n", category="N"))
    xlsx_path = tmp_path / "export.xlsx"

    export_to_excel(repo, xlsx_path)

    workbook = load_workbook(xlsx_path)
    assert workbook["AM — Prioritaire"].max_row == 2  # header + 1 row
    assert workbook["Finance — Autres opportunités"].max_row == 2
    assert workbook["Hors sujet"].max_row == 2


def test_export_includes_score_and_score_detail_columns(tmp_path):
    repo = SQLiteJobRepository(tmp_path / "postings.db")
    repo.add(make_posting(category="A", location_priority=1))
    xlsx_path = tmp_path / "export.xlsx"

    export_to_excel(repo, xlsx_path)

    workbook = load_workbook(xlsx_path)
    sheet = workbook["AM — Prioritaire"]
    headers = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    assert "Score" in headers
    assert "Score - Détail" in headers
    row = dict(zip(headers, next(sheet.iter_rows(min_row=2, max_row=2, values_only=True))))
    assert isinstance(row["Score"], int)
    assert row["Score"] > 0
    assert row["Score - Détail"]


def test_export_sorts_postings_by_score_descending_within_a_sheet(tmp_path):
    repo = SQLiteJobRepository(tmp_path / "postings.db")
    # Lower-scoring posting (farther location) added first, higher-scoring
    # one added second — the export must still put the higher score first.
    repo.add(make_posting(url="https://example.com/far", title="Far posting", category="A", location_priority=4))
    repo.add(make_posting(url="https://example.com/near", title="Near posting", category="A", location_priority=1))
    xlsx_path = tmp_path / "export.xlsx"

    export_to_excel(repo, xlsx_path)

    workbook = load_workbook(xlsx_path)
    sheet = workbook["AM — Prioritaire"]
    headers = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    rows = [dict(zip(headers, r)) for r in sheet.iter_rows(min_row=2, values_only=True)]

    assert [r["Titre du poste"] for r in rows] == ["Near posting", "Far posting"]


def test_export_includes_reason_location_and_letter_columns(tmp_path):
    repo = SQLiteJobRepository(tmp_path / "postings.db")
    posting = make_posting(
        category="A",
        classification_reason="Analyse de portefeuille",
        location="Paris, France",
        location_priority=1,
    )
    repo.add(posting)
    repo.update_letter(posting.stable_id(), "Corps de la lettre...")
    xlsx_path = tmp_path / "export.xlsx"

    export_to_excel(repo, xlsx_path)

    workbook = load_workbook(xlsx_path)
    sheet = workbook["AM — Prioritaire"]
    headers = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    row = dict(zip(headers, next(sheet.iter_rows(min_row=2, max_row=2, values_only=True))))
    assert row["Raison (classification)"] == "Analyse de portefeuille"
    assert row["Lieu"] == "Paris, France"
    assert row["Priorité géo"] == 1
    assert row["Lettre de motivation générée"] == "Corps de la lettre..."


def test_export_overwrites_previous_snapshot_rather_than_appending(tmp_path):
    repo = SQLiteJobRepository(tmp_path / "postings.db")
    repo.add(make_posting(url="https://example.com/a", category="A"))
    xlsx_path = tmp_path / "export.xlsx"

    export_to_excel(repo, xlsx_path)
    export_to_excel(repo, xlsx_path)

    workbook = load_workbook(xlsx_path)
    assert workbook["AM — Prioritaire"].max_row == 2
